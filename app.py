from flask import Flask, render_template, request, jsonify, send_file
from datetime import datetime
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A3, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
import random

app = Flask(__name__)

# In-memory storage
timetables = {}
# Short-term session memory: keyed by session_id, resets with process
session_memory = {}

class TimetableGenerator:
    def __init__(self, teachers, subjects, classrooms, time_slots, days, semesters):
        self.teachers = teachers
        self.subjects = subjects
        self.classrooms = classrooms
        self.time_slots = time_slots
        self.days = days
        self.semesters = semesters
        self.timetable = []
        self.conflicts = []
        

        
    def _subject_color(self, name: str) -> str:
        # Deterministic HSL pastel color from subject name
        h = 0
        for ch in name:
            h = (h * 31 + ord(ch)) % 360
        return f"hsl({h},65%,85%)"

    def _rank_slots(self, cohort_load, day, slot, cohort):
        # Lower score is better
        day_load = cohort_load.get(cohort, {}).get(day, 0)
        # prefer days with lower load and contiguous slots within a day
        contiguous_bonus = 0
        # in absence of exact time order semantics, no adjacency detection beyond load
        score = day_load * 10 - contiguous_bonus
        return score

    def generate(self):
        """Generate timetable with conflict detection, optimization, and complete scheduling"""
        self.timetable = []
        self.conflicts = []
        
        # Initialize schedules for tracking
        teacher_schedule = {
            teacher['name']: {
                day: {slot: False for slot in self.time_slots} 
                for day in self.days
            } 
            for teacher in self.teachers
        }
        
        classroom_schedule = {
            classroom: {
                day: {slot: False for slot in self.time_slots} 
                for day in self.days
            } 
            for classroom in self.classrooms
        }

        # cohort schedule by semester
        cohort_schedule = {
            semester: {
                day: {slot: False for slot in self.time_slots}
                for day in self.days
            }
            for semester in (set([s.get('semester','General') for s in self.subjects]) or {'General'})
        }

        # cohort daily load counts
        cohort_load = {}

        def can_place(subject, teacher, classroom, day, slot):
            semester = subject.get('semester', 'General')
            if teacher_schedule[teacher['name']][day][slot]:
                return False
            if classroom_schedule[classroom][day][slot]:
                return False
            if cohort_schedule[semester][day][slot]:
                return False
            # teacher available?
            if not self.is_teacher_available(teacher, day, slot):
                return False
            return True

        def place_entry(subject, teacher, classroom, day, slot):
            semester = subject.get('semester', 'General')
            entry = {
                'day': day,
                'time_slot': slot,
                'subject': subject['name'],
                'teacher': teacher['name'],
                'classroom': classroom,
                'semester': semester,
                'subject_color': self._subject_color(subject['name'])
            }
            self.timetable.append(entry)
            teacher_schedule[teacher['name']][day][slot] = True
            classroom_schedule[classroom][day][slot] = True
            cohort_schedule[semester][day][slot] = True
            # load increment
            cohort_load.setdefault(semester, {}).setdefault(day, 0)
            cohort_load[semester][day] += 1

        def suggest_alternatives(subject, max_suggestions=5):
            suggestions = []
            # iterate all day/slot for free positions
            for day in self.days:
                for slot in self.time_slots:
                    # available teachers for subject
                    available_teachers = [
                        t for t in self.teachers
                        if subject['name'].lower() in [s.lower() for s in t.get('subjects', [])] and
                           not teacher_schedule[t['name']][day][slot] and
                           self.is_teacher_available(t, day, slot)
                    ]
                    if not available_teachers:
                        continue
                    # available classrooms
                    available_classrooms = [c for c in self.classrooms if not classroom_schedule[c][day][slot]]
                    if not available_classrooms:
                        continue
                    semester = subject.get('semester','General')
                    if cohort_schedule[semester][day][slot]:
                        continue
                    # feasible
                    suggestions.append(f"{day} @ {slot}")
                    if len(suggestions) >= max_suggestions:
                        return suggestions
            return suggestions

        # Assign classes per subject with ranked slots
        for subject in self.subjects:
            sessions_per_week = subject.get('sessions_per_week', 2)
            placed = 0
            attempts_guard = 0
            while placed < sessions_per_week and attempts_guard < sessions_per_week * 300:
                attempts_guard += 1
                # candidate list of (day, slot) sorted by rank
                semester = subject.get('semester', 'General')
                candidates = []
                for day in self.days:
                    for slot in self.time_slots:
                        score = self._rank_slots(cohort_load, day, slot, semester)
                        candidates.append((score, day, slot))
                candidates.sort(key=lambda x: x[0])

                assigned_here = False
                for _, day, slot in candidates:
                    available_teachers = [
                        t for t in self.teachers
                        if subject['name'].lower() in [s.lower() for s in t.get('subjects', [])]
                    ]
                    random.shuffle(available_teachers)
                    for teacher in available_teachers:
                        if not self.is_teacher_available(teacher, day, slot):
                            continue
                        # find a free classroom
                        free_classrooms = [c for c in self.classrooms if not classroom_schedule[c][day][slot]]
                        random.shuffle(free_classrooms)
                        for classroom in free_classrooms:
                            if can_place(subject, teacher, classroom, day, slot):
                                place_entry(subject, teacher, classroom, day, slot)
                                placed += 1
                                assigned_here = True
                                break
                        if assigned_here:
                            break
                    if assigned_here:
                        break

                if not assigned_here:
                    # could not place this session now, break to avoid infinite loop
                    break

            # If not fully placed, record conflicts with suggestions
            if placed < sessions_per_week:
                missing = sessions_per_week - placed
                suggestions = suggest_alternatives(subject)
                self.conflicts.append({
                    'type': 'student',
                    'semester': subject.get('semester','General'),
                    'time_slot': None,
                    'day': None,
                    'subjects': [subject['name']],
                    'missing_sessions': missing,
                    'suggestions': suggestions
                })
        
        # Detect conflicts (double safety)
        self.detect_conflicts()
        
        return {
            'timetable': self.timetable,
            'conflicts': self.conflicts
        }
    
    def is_teacher_available(self, teacher, day, time_slot):
        """Check if teacher is available at given time"""
        availability = teacher.get('availability', {})
        if day not in availability:
            return True
        return time_slot in availability[day]
    
    def detect_conflicts(self):
        """Detect scheduling conflicts across teacher/classroom/student and attach suggestions"""
        conflicts = []
        indexed = {}
        for e in self.timetable:
            key = (e['day'], e['time_slot'])
            indexed.setdefault(key, []).append(e)

        for (day, slot), entries in indexed.items():
            # teacher conflicts
            teacher_map = {}
            classroom_map = {}
            cohort_map = {}
            for e in entries:
                teacher_map.setdefault(e['teacher'], []).append(e)
                classroom_map.setdefault(e['classroom'], []).append(e)
                cohort_map.setdefault(e.get('semester','General'), []).append(e)
            for t, arr in teacher_map.items():
                if len(arr) > 1:
                    conflicts.append({
                        'type': 'teacher', 'teacher': t, 'day': day, 'time_slot': slot,
                        'subjects': [x['subject'] for x in arr]
                    })
            for c, arr in classroom_map.items():
                if len(arr) > 1:
                    conflicts.append({
                        'type': 'classroom', 'classroom': c, 'day': day, 'time_slot': slot,
                        'subjects': [x['subject'] for x in arr]
                    })
            for cohort, arr in cohort_map.items():
                if len(arr) > 1:
                    conflicts.append({
                        'type': 'student', 'semester': cohort, 'day': day, 'time_slot': slot,
                        'subjects': [x['subject'] for x in arr]
                    })
        self.conflicts.extend(conflicts)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/generate', methods=['POST'])
def generate_timetable():
    try:
        data = request.json
        # Optional existing session_id to continue a session
        existing_session_id = data.get('session_id')
        
        # Pull from memory if fields are missing and a session exists
        memory = session_memory.get(existing_session_id, {}) if existing_session_id else {}
        
        teachers = data.get('teachers', memory.get('teachers', []))
        subjects = data.get('subjects', memory.get('subjects', []))
        classrooms = data.get('classrooms', memory.get('classrooms', []))
        time_slots = data.get('timeSlots', memory.get('timeSlots', []))
        days = data.get('days', memory.get('days', []))
        semesters = data.get('semesters', memory.get('semesters', []))
        preferences = data.get('preferences', memory.get('preferences', {}))
        
        if not all([teachers, subjects, classrooms, time_slots, days]):
            return jsonify({'error': 'Missing required data'}), 400
        
        generator = TimetableGenerator(teachers, subjects, classrooms, time_slots, days, semesters)
        result = generator.generate()
        
        # Decide session_id: reuse if provided, else create new
        import secrets
        session_id = existing_session_id or secrets.token_hex(8)
        timetables[session_id] = {
            'timetable': result['timetable'],
            'conflicts': result['conflicts'],
            'metadata': {
                'classrooms': classrooms,
                'days': days,
                'time_slots': time_slots,
                'semesters': semesters
            }
        }
        
        # Update short-term memory for this session
        session_memory[session_id] = {
            'teachers': teachers,
            'subjects': subjects,
            'classrooms': classrooms,
            'timeSlots': time_slots,
            'days': days,
            'semesters': semesters,
            'preferences': preferences,
            'last_updated': datetime.now().isoformat()
        }
        
        return jsonify({
            'success': True,
            'session_id': session_id,
            'timetable': result['timetable'],
            'conflicts': result['conflicts'],
            'memory': session_memory[session_id]
        })
        
    except Exception as e:
        print(f"Error: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/update-timetable', methods=['POST'])
def update_timetable():
    try:
        data = request.json
        session_id = data.get('session_id')
        timetable = data.get('timetable')
        # Optional updates to memory fields to let user modify previous inputs
        mem_updates = data.get('memory_updates', {})
        
        if session_id in timetables:
            timetables[session_id]['timetable'] = timetable
            
            # Re-detect conflicts
            generator = TimetableGenerator([], [], [], [], [], [])
            generator.timetable = timetable
            generator.detect_conflicts()
            timetables[session_id]['conflicts'] = generator.conflicts

            # Apply memory updates if provided
            if session_id in session_memory and isinstance(mem_updates, dict):
                session_memory[session_id].update({k: v for k, v in mem_updates.items() if v is not None})
                session_memory[session_id]['last_updated'] = datetime.now().isoformat()
            
            return jsonify({
                'success': True,
                'conflicts': generator.conflicts,
                'memory': session_memory.get(session_id)
            })
        
        return jsonify({'error': 'Session not found'}), 404
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/export/excel/<session_id>')
def export_excel(session_id):
    try:
        if session_id not in timetables:
            return "Timetable not found", 404
        
        timetable_data = timetables[session_id]['timetable']
        metadata = timetables[session_id]['metadata']
        
        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)
        
        classrooms = metadata['classrooms']
        days = metadata['days']
        time_slots = metadata['time_slots']
        
        # Define styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        day_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        day_font = Font(bold=True, size=10)
        break_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
        break_font = Font(bold=True, color="856404", size=10)
        class_fill = PatternFill(start_color="E7F3E7", end_color="E7F3E7", fill_type="solid")
        title_font = Font(bold=True, size=14, color="4472C4")
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Create a single sheet combining all classrooms
        ws = wb.create_sheet(title='Unified Timetable')

        # Title
        ws['A1'] = 'Unified Weekly Timetable'
        ws['A1'].font = title_font
        # Columns: Day + for each time slot, one column containing concatenated classroom entries
        total_cols = 1 + len(time_slots)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

        # Column headers (time slots)
        ws['A2'] = 'Day / Time'
        ws['A2'].fill = header_fill
        ws['A2'].font = header_font
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A2'].border = thin_border

        for col, slot in enumerate(time_slots, start=2):
            cell = ws.cell(row=2, column=col)
            cell.value = slot
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border

        # Fill data: for each day and slot, list all classroom entries combined
        for row, day in enumerate(days, start=3):
            day_cell = ws.cell(row=row, column=1)
            day_cell.value = day
            day_cell.fill = day_fill
            day_cell.font = day_font
            day_cell.alignment = Alignment(horizontal='center', vertical='center')
            day_cell.border = thin_border

            for col, slot in enumerate(time_slots, start=2):
                # Collect entries for all classrooms at this day/slot
                entries = [
                    e for e in timetable_data
                    if e['day'] == day and e['time_slot'] == slot
                ]
                cell = ws.cell(row=row, column=col)
                if entries:
                    # Compose multiline text: Classroom: Subject (Semester) - Teacher
                    lines = [
                        f"{e['classroom']}: {e['subject']}\n{e['teacher']} ({e.get('semester', '-')})"
                        for e in sorted(entries, key=lambda x: x['classroom'])
                    ]
                    cell.value = "\n\n".join(lines)
                    cell.fill = class_fill
                    cell.font = Font(size=9)
                else:
                    cell.value = ""
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border

        # Adjust dimensions
        ws.column_dimensions['A'].width = 15
        for col in range(2, len(time_slots) + 2):
            ws.column_dimensions[chr(64 + col)].width = 28

        ws.row_dimensions[1].height = 25
        ws.row_dimensions[2].height = 30
        for row in range(3, len(days) + 3):
            ws.row_dimensions[row].height = 90
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'timetable_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return str(e), 500

# Memory utility endpoints for frontend integration
@app.route('/memory/<session_id>', methods=['GET'])
def get_memory(session_id):
    if session_id in session_memory:
        return jsonify({'success': True, 'memory': session_memory[session_id]})
    return jsonify({'error': 'Session not found'}), 404

@app.route('/memory/<session_id>', methods=['POST'])
def update_memory(session_id):
    data = request.json or {}
    if session_id not in session_memory:
        session_memory[session_id] = {}
    session_memory[session_id].update({k: v for k, v in data.items() if v is not None})
    session_memory[session_id]['last_updated'] = datetime.now().isoformat()
    return jsonify({'success': True, 'memory': session_memory[session_id]})

@app.route('/memory/clear/<session_id>', methods=['POST'])
def clear_memory(session_id):
    if session_id in session_memory:
        del session_memory[session_id]
        return jsonify({'success': True})
    return jsonify({'error': 'Session not found'}), 404

@app.route('/export/pdf/<session_id>')
def export_pdf(session_id):
    try:
        if session_id not in timetables:
            return "Timetable not found", 404
        
        timetable_data = timetables[session_id]['timetable']
        metadata = timetables[session_id]['metadata']
        conflicts = timetables[session_id].get('conflicts', [])
        
        # Create PDF
        output = io.BytesIO()
        doc = SimpleDocTemplate(
            output, 
            pagesize=landscape(A3), 
            topMargin=0.5*inch, 
            bottomMargin=0.5*inch,
            leftMargin=0.5*inch,
            rightMargin=0.5*inch
        )
        elements = []
        
        styles = getSampleStyleSheet()
        classrooms = metadata['classrooms']
        days = metadata['days']
        time_slots = metadata['time_slots']
        
        # Create a single unified timetable across all classrooms
        title = Paragraph("<b>Unified Weekly Timetable</b>", styles['Title'])
        elements.append(title)
        elements.append(Spacer(1, 0.2 * inch))

        # Create table data: Day/Time + for each slot, one cell containing combined entries for all classrooms
        table_data = [['Day / Time'] + time_slots]

        for day in days:
            row = [day]
            for slot in time_slots:
                entries = [e for e in timetable_data if e['day'] == day and e['time_slot'] == slot]
                if entries:
                    # rich text per entry, but background per cell later; include classroom and semester
                    lines = [
                        f"{e['classroom']}: {e['subject']}\n{e['teacher']} ({e.get('semester', '-')})"
                        for e in sorted(entries, key=lambda x: x['classroom'])
                    ]
                    row.append("\n\n".join(lines))
                else:
                    row.append('')
            table_data.append(row)

        # Compute column widths to span full page width
        available_width = doc.width
        n_time_cols = len(time_slots)
        day_col_width = max(72, available_width * 0.12)
        remaining_width = max(0, available_width - day_col_width)
        slot_col_width = remaining_width / n_time_cols if n_time_cols > 0 else remaining_width
        col_widths = [day_col_width] + [slot_col_width] * n_time_cols

        # Create and style table with enhanced design
        table = Table(table_data, repeatRows=1, colWidths=col_widths)

        header_bg = colors.Color(0.18, 0.36, 0.6)
        header_text = colors.whitesmoke
        day_col_bg = colors.HexColor('#E9EFF8')
        grid_color = colors.HexColor('#B0BEC5')
        zebra_bg = colors.HexColor('#F7FAFC')

        style_list = [
            ('BACKGROUND', (0, 0), (-1, 0), header_bg),
            ('TEXTCOLOR', (0, 0), (-1, 0), header_text),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('TOPPADDING', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, 1), (0, -1), day_col_bg),
            ('TEXTCOLOR', (0, 1), (0, -1), colors.HexColor('#0F3057')),
            ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 1), (0, -1), 10),
            ('FONTNAME', (1, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (1, 1), (-1, -1), 9),
            ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, grid_color),
        ]

        # Zebra striping for rows (excluding header)
        for r in range(1, len(table_data)):
            if r % 2 == 0:
                style_list.append(('BACKGROUND', (1, r), (-1, r), zebra_bg))

        # Per-cell subject color is not trivial in a combined multi-entry cell; keep readable base and subtle highlight if any entries exist
        for row_idx, day in enumerate(days, start=1):
            for col_idx, slot in enumerate(time_slots, start=1):
                has_any = any(e for e in timetable_data if e['day'] == day and e['time_slot'] == slot)
                if has_any:
                    style_list.append(('BACKGROUND', (col_idx, row_idx), (col_idx, row_idx), colors.HexColor('#E8F5E9')))

        table.setStyle(TableStyle(style_list))
        elements.append(table)

        # Conflicts page
        if conflicts:
            elements.append(PageBreak())
            elements.append(Paragraph("<b>Conflicts</b>", styles['Title']))
            elements.append(Spacer(1, 0.2*inch))

            conflict_headers = ['Type','Day','Time','Teacher','Classroom','Semester','Subjects','Suggestions']
            conflict_rows = [conflict_headers]
            type_map = {'teacher':'Teacher','classroom':'Classroom','student':'Student'}

            for c in conflicts:
                conflict_rows.append([
                    type_map.get(c.get('type'), 'Conflict'),
                    c.get('day','-') or '-',
                    c.get('time_slot','-') or '-',
                    c.get('teacher','-') or '-',
                    c.get('classroom','-') or '-',
                    c.get('semester','-') or '-',
                    ", ".join(c.get('subjects', []) or []),
                    ", ".join(c.get('suggestions', []) or [])
                ])
            conflicts_table = Table(conflict_rows, repeatRows=1, colWidths=[80,60,60,110,110,90,200,200])
            conflicts_table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), header_bg),
                ('TEXTCOLOR', (0,0), (-1,0), header_text),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('ALIGN', (0,0), (-1,-1), 'LEFT'),
                ('GRID', (0,0), (-1,-1), 0.5, grid_color),
                ('FONTSIZE', (0,0), (-1,0), 11),
                ('FONTSIZE', (0,1), (-1,-1), 8),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (-1, -1), 6),
                ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ]))
            elements.append(conflicts_table)
        
        doc.build(elements)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'timetable_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        )
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return str(e), 500

@app.route('/conflicts/<session_id>', methods=['GET'])
def get_conflicts(session_id):
    if session_id not in timetables:
        return jsonify({'error': 'Session not found'}), 404
    return jsonify({'success': True, 'conflicts': timetables[session_id].get('conflicts', [])})

if __name__ == '__main__':
    app.run(debug=True, port=5000)